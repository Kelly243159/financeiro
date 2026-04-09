import streamlit as st
import pandas as pd
import re
import os
import hashlib
import json
import calendar
from copy import copy
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import tempfile

PARCELA_REGEX = re.compile(r"\((\d+)/(\d+)\)")

MESES_PT = {
    1: "Janeiro", 2: "Fevereiro", 3: "Março", 4: "Abril",
    5: "Maio", 6: "Junho", 7: "Julho", 8: "Agosto",
    9: "Setembro", 10: "Outubro", 11: "Novembro", 12: "Dezembro",
}

MESES_PT_UPPER = {
    "JANEIRO": 1, "FEVEREIRO": 2, "MARÇO": 3, "MARCO": 3, "ABRIL": 4,
    "MAIO": 5, "JUNHO": 6, "JULHO": 7, "AGOSTO": 8,
    "SETEMBRO": 9, "OUTUBRO": 10, "NOVEMBRO": 11, "DEZEMBRO": 12,
}

COR_HEADER_BG = "0F1B2D"
COR_HEADER_FONT = "FFFFFF"
COR_TITULO_BG = "1B3A5C"
COR_TITULO_FONT = "FFFFFF"
COR_ZEBRA_A = "FFFFFF"
COR_ZEBRA_B = "EDF2F7"
COR_RESUMO_BG = "E2E8F0"
COR_RESUMO_FONT = "0F1B2D"
COR_BORDA = "CBD5E1"

# ═══════════════════════════════════════════
# CONFIGURAÇÃO DE USUÁRIOS
# ═══════════════════════════════════════════

def _hash_senha(senha):
    return hashlib.sha256(senha.encode()).hexdigest()

USUARIOS = {
    "mvtec2026": {
        "senha_hash": _hash_senha("MV@@2026"),
        "nome": "MV TEC",
        "perfil": "admin",
    },
}


def autenticar(usuario, senha):
    """Verifica credenciais e retorna dados do usuário ou None."""
    user_data = USUARIOS.get(usuario.lower().strip())
    if user_data and user_data["senha_hash"] == _hash_senha(senha):
        return user_data
    return None


# ═══════════════════════════════════════════
# TELA DE LOGIN
# ═══════════════════════════════════════════

def tela_login():
    """Renderiza a tela de login com estilo profissional."""

    st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=DM+Sans:ital,opsz,wght@0,9..40,300;0,9..40,400;0,9..40,500;0,9..40,700;1,9..40,400&display=swap');

    :root {
        --mv-navy:    #0f1b2d;
        --mv-blue:    #1b3a5c;
        --mv-accent:  #3b82f6;
        --mv-accent2: #60a5fa;
        --mv-surface: #f8fafc;
        --mv-border:  #e2e8f0;
        --mv-text:    #1e293b;
        --mv-muted:   #64748b;
        --mv-radius:  14px;
    }

    .stApp {
        background: linear-gradient(160deg, #0f1b2d 0%, #1b3a5c 40%, #234e7a 70%, #2d6498 100%) !important;
        min-height: 100vh;
    }

    html, body, .stApp, .stApp * {
        font-family: 'DM Sans', sans-serif !important;
    }

    /* ── Container do login ── */
    .login-wrapper {
        display: flex;
        flex-direction: column;
        align-items: center;
        justify-content: center;
        padding: 2rem 1rem;
        min-height: 70vh;
    }

    .login-card {
        background: rgba(255, 255, 255, 0.97);
        border-radius: 20px;
        padding: 2.8rem 2.4rem 2.2rem;
        width: 100%;
        max-width: 400px;
        box-shadow:
            0 4px 6px rgba(0,0,0,0.07),
            0 20px 60px rgba(0,0,0,0.15),
            0 0 0 1px rgba(255,255,255,0.1);
        position: relative;
        overflow: hidden;
    }

    .login-card::before {
        content: '';
        position: absolute;
        top: 0; left: 0; right: 0;
        height: 4px;
        background: linear-gradient(90deg, #3b82f6, #60a5fa, #3b82f6);
        background-size: 200% 100%;
        animation: shimmer 3s ease infinite;
    }

    @keyframes shimmer {
        0%, 100% { background-position: 0% 50%; }
        50% { background-position: 100% 50%; }
    }

    .login-logo {
        text-align: center;
        margin-bottom: 0.3rem;
    }

    .login-logo .icon-circle {
        width: 64px; height: 64px;
        background: linear-gradient(135deg, #0f1b2d, #1b3a5c);
        border-radius: 16px;
        display: inline-flex;
        align-items: center;
        justify-content: center;
        font-size: 1.8rem;
        margin-bottom: 0.8rem;
        box-shadow: 0 4px 12px rgba(15,27,45,0.2);
    }

    .login-logo h2 {
        color: #0f1b2d;
        font-size: 1.25rem;
        font-weight: 700;
        letter-spacing: 2.5px;
        margin: 0;
    }

    .login-logo p {
        color: #64748b;
        font-size: 0.78rem;
        margin: 0.25rem 0 0;
        letter-spacing: 0.3px;
    }

    .login-divider {
        height: 1px;
        background: linear-gradient(90deg, transparent, #e2e8f0, transparent);
        margin: 1.4rem 0 1.6rem;
    }

    /* ── Estilizar inputs ── */
    .login-card [data-testid="stTextInput"] label,
    .login-card [data-testid="stPasswordInput"] label {
        font-size: 0.8rem !important;
        font-weight: 600 !important;
        color: #1e293b !important;
        letter-spacing: 0.5px;
        text-transform: uppercase;
    }

    .login-card [data-testid="stTextInput"] input,
    .login-card [data-testid="stPasswordInput"] input {
        border-radius: 10px !important;
        border: 1.5px solid #e2e8f0 !important;
        padding: 0.6rem 0.9rem !important;
        font-size: 0.92rem !important;
        transition: all 0.2s ease !important;
        background: #f8fafc !important;
    }

    .login-card [data-testid="stTextInput"] input:focus,
    .login-card [data-testid="stPasswordInput"] input:focus {
        border-color: #3b82f6 !important;
        box-shadow: 0 0 0 3px rgba(59,130,246,0.1) !important;
        background: #fff !important;
    }

    /* ── Botão de login ── */
    .login-card .stButton > button[kind="primary"],
    .login-card button[data-testid="stBaseButton-primary"] {
        background: linear-gradient(135deg, #0f1b2d, #1b3a5c) !important;
        color: #fff !important;
        border: none !important;
        border-radius: 10px !important;
        padding: 0.7rem 1.5rem !important;
        font-weight: 600 !important;
        font-size: 0.92rem !important;
        letter-spacing: 0.5px;
        transition: all 0.25s ease;
        margin-top: 0.5rem;
    }

    .login-card .stButton > button[kind="primary"]:hover,
    .login-card button[data-testid="stBaseButton-primary"]:hover {
        background: linear-gradient(135deg, #1b3a5c, #234e7a) !important;
        box-shadow: 0 6px 20px rgba(15,27,45,0.3);
        transform: translateY(-1px);
    }

    .login-footer {
        text-align: center;
        margin-top: 1.2rem;
        color: #94a3b8;
        font-size: 0.72rem;
        letter-spacing: 0.3px;
    }

    /* ── Erro de login ── */
    .login-error {
        background: rgba(239,68,68,0.08);
        border: 1px solid rgba(239,68,68,0.2);
        border-radius: 10px;
        padding: 0.6rem 1rem;
        color: #dc2626;
        font-size: 0.82rem;
        font-weight: 500;
        text-align: center;
        margin-top: 0.5rem;
        animation: shakeX 0.5s ease;
    }

    @keyframes shakeX {
        0%, 100% { transform: translateX(0); }
        20%, 60% { transform: translateX(-6px); }
        40%, 80% { transform: translateX(6px); }
    }

    /* ── Partículas de fundo ── */
    .bg-particles {
        position: fixed;
        top: 0; left: 0;
        width: 100%; height: 100%;
        pointer-events: none;
        overflow: hidden;
        z-index: 0;
    }

    .bg-particles .orb {
        position: absolute;
        border-radius: 50%;
        opacity: 0.06;
        background: #60a5fa;
        animation: floatOrb 20s ease-in-out infinite;
    }

    .bg-particles .orb:nth-child(1) {
        width: 300px; height: 300px; top: 10%; left: -5%;
        animation-delay: 0s;
    }
    .bg-particles .orb:nth-child(2) {
        width: 200px; height: 200px; top: 60%; right: -3%;
        animation-delay: -7s;
    }
    .bg-particles .orb:nth-child(3) {
        width: 150px; height: 150px; bottom: 10%; left: 30%;
        animation-delay: -14s;
    }

    @keyframes floatOrb {
        0%, 100% { transform: translate(0, 0) scale(1); }
        33% { transform: translate(30px, -20px) scale(1.05); }
        66% { transform: translate(-20px, 15px) scale(0.95); }
    }

    #MainMenu, footer, header { visibility: hidden; }
    </style>

    <div class="bg-particles">
        <div class="orb"></div>
        <div class="orb"></div>
        <div class="orb"></div>
    </div>
    """, unsafe_allow_html=True)

    # Layout centralizado
    col_left, col_center, col_right = st.columns([1, 2, 1])

    with col_center:
        st.markdown("""
        <div class="login-wrapper">
            <div class="login-card">
                <div class="login-logo">
                    <div class="icon-circle">📊</div>
                    <h2>MV TEC</h2>
                    <p>Gestão de parcelas e empréstimos</p>
                </div>
                <div class="login-divider"></div>
        """, unsafe_allow_html=True)

        usuario = st.text_input("Usuário", placeholder="Digite seu usuário", key="login_user")
        senha = st.text_input("Senha", type="password", placeholder="Digite sua senha", key="login_pass")

        if st.button("Entrar", use_container_width=True, type="primary", key="login_btn"):
            if usuario and senha:
                user_data = autenticar(usuario, senha)
                if user_data:
                    st.session_state["autenticado"] = True
                    st.session_state["usuario"] = usuario
                    st.session_state["nome_usuario"] = user_data["nome"]
                    st.session_state["perfil"] = user_data["perfil"]
                    st.rerun()
                else:
                    st.markdown(
                        '<div class="login-error">⚠ Usuário ou senha incorretos</div>',
                        unsafe_allow_html=True,
                    )
            else:
                st.markdown(
                    '<div class="login-error">Preencha todos os campos</div>',
                    unsafe_allow_html=True,
                )

        st.markdown("""
                <div class="login-footer">
                    © 2026 MV TEC — Acesso restrito
                </div>
            </div>
        </div>
        """, unsafe_allow_html=True)


# ═══════════════════════════════════════════
# GERADOR DE RELATÓRIO PROFISSIONAL
# ═══════════════════════════════════════════

def _gerar_relatorio(ws_origem, linhas_quitadas, mes, ano):
    wb_rel = Workbook()
    ws = wb_rel.active
    ws.title = "Quitados"

    total_colunas_orig = ws_origem.max_column
    borda_fina = Border(
        left=Side(style="thin", color=COR_BORDA),
        right=Side(style="thin", color=COR_BORDA),
        top=Side(style="thin", color=COR_BORDA),
        bottom=Side(style="thin", color=COR_BORDA),
    )

    colunas_com_dados = []
    for col in range(1, total_colunas_orig + 1):
        header_val = ws_origem.cell(row=1, column=col).value
        if header_val is None or str(header_val).strip() == "":
            continue
        tem_dado = False
        for row in linhas_quitadas:
            val = ws_origem.cell(row=row, column=col).value
            if val is not None and str(val).strip() != "":
                tem_dado = True
                break
        if tem_dado:
            colunas_com_dados.append(col)

    if not colunas_com_dados and linhas_quitadas:
        for col in range(1, total_colunas_orig + 1):
            header_val = ws_origem.cell(row=1, column=col).value
            if header_val is not None and str(header_val).strip() != "":
                colunas_com_dados.append(col)

    total_colunas = len(colunas_com_dados)
    if total_colunas == 0:
        total_colunas = total_colunas_orig
        colunas_com_dados = list(range(1, total_colunas_orig + 1))

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_colunas)
    titulo_cell = ws.cell(row=1, column=1)
    nome_mes = MESES_PT.get(int(mes), mes)
    titulo_cell.value = f"RELATÓRIO DE PARCELAS QUITADAS — {nome_mes.upper()} / {ano}"
    titulo_cell.font = Font(name="Arial", bold=True, size=13, color=COR_TITULO_FONT)
    titulo_cell.fill = PatternFill("solid", fgColor=COR_TITULO_BG)
    titulo_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36
    for c in range(2, total_colunas + 1):
        ws.cell(row=1, column=c).fill = PatternFill("solid", fgColor=COR_TITULO_BG)

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_colunas)
    sub_cell = ws.cell(row=2, column=1)
    sub_cell.value = f"Gerado em {datetime.now().strftime('%d/%m/%Y às %H:%M')}  •  Total de registros: {len(linhas_quitadas)}"
    sub_cell.font = Font(name="Arial", size=9, color="64748B", italic=True)
    sub_cell.fill = PatternFill("solid", fgColor=COR_ZEBRA_B)
    sub_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 22
    for c in range(2, total_colunas + 1):
        ws.cell(row=2, column=c).fill = PatternFill("solid", fgColor=COR_ZEBRA_B)

    header_fill = PatternFill("solid", fgColor=COR_HEADER_BG)
    header_font = Font(name="Arial", bold=True, size=10, color=COR_HEADER_FONT)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for dest_col, orig_col in enumerate(colunas_com_dados, start=1):
        cell = ws.cell(row=3, column=dest_col, value=ws_origem.cell(row=1, column=orig_col).value)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = borda_fina
    ws.row_dimensions[3].height = 28

    fill_a = PatternFill("solid", fgColor=COR_ZEBRA_A)
    fill_b = PatternFill("solid", fgColor=COR_ZEBRA_B)
    data_font = Font(name="Arial", size=10, color="1E293B")
    data_align = Alignment(vertical="center", wrap_text=True)
    data_align_center = Alignment(horizontal="center", vertical="center")

    linha_destino = 4
    for idx, row in enumerate(linhas_quitadas):
        fill = fill_a if idx % 2 == 0 else fill_b
        for dest_col, orig_col in enumerate(colunas_com_dados, start=1):
            valor = ws_origem.cell(row=row, column=orig_col).value
            cell = ws.cell(row=linha_destino, column=dest_col, value=valor)
            cell.font = data_font
            cell.fill = fill
            cell.border = borda_fina
            if isinstance(valor, (int, float)):
                cell.alignment = data_align_center
                if isinstance(valor, float):
                    cell.number_format = '#,##0.00'
                elif isinstance(valor, int) and valor > 100:
                    cell.number_format = '#,##0'
            else:
                cell.alignment = data_align
        ws.row_dimensions[linha_destino].height = 22
        linha_destino += 1

    if linhas_quitadas:
        resumo_row = linha_destino + 1
        resumo_fill = PatternFill("solid", fgColor=COR_RESUMO_BG)
        resumo_font = Font(name="Arial", bold=True, size=10, color=COR_RESUMO_FONT)

        merge_end = min(2, total_colunas)
        if merge_end > 1:
            ws.merge_cells(start_row=resumo_row, start_column=1, end_row=resumo_row, end_column=merge_end)
        cell_label = ws.cell(row=resumo_row, column=1, value="TOTAL QUITADOS")
        cell_label.font = resumo_font
        cell_label.fill = resumo_fill
        cell_label.alignment = Alignment(horizontal="right", vertical="center")
        cell_label.border = borda_fina
        if merge_end > 1:
            ws.cell(row=resumo_row, column=2).fill = resumo_fill
            ws.cell(row=resumo_row, column=2).border = borda_fina

        val_col = min(3, total_colunas)
        if val_col > merge_end:
            cell_val = ws.cell(row=resumo_row, column=val_col, value=len(linhas_quitadas))
            cell_val.font = Font(name="Arial", bold=True, size=12, color=COR_RESUMO_FONT)
            cell_val.fill = resumo_fill
            cell_val.alignment = Alignment(horizontal="center", vertical="center")
            cell_val.border = borda_fina

        for col in range(val_col + 1, total_colunas + 1):
            c = ws.cell(row=resumo_row, column=col)
            c.fill = resumo_fill
            c.border = borda_fina
        ws.row_dimensions[resumo_row].height = 28

    for dest_col in range(1, total_colunas + 1):
        max_len = 0
        col_letter = get_column_letter(dest_col)
        for row_idx in range(1, linha_destino + 2):
            val = ws.cell(row=row_idx, column=dest_col).value
            if val:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 45)

    ws.freeze_panes = "A4"

    if linhas_quitadas:
        last_data_row = 3 + len(linhas_quitadas)
        ws.auto_filter.ref = f"A3:{get_column_letter(total_colunas)}{last_data_row}"

    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:3"

    return wb_rel


# ═══════════════════════════════════════════
# FUNÇÕES DE PROCESSAMENTO (PARCELAS)
# ═══════════════════════════════════════════

def processar_planilha(caminho_entrada, mes, ano):
    wb = load_workbook(caminho_entrada)
    ws = wb.active
    linhas_quitadas = []

    for row in range(2, ws.max_row + 1):
        cell = ws[f"H{row}"]
        valor = cell.value
        if not isinstance(valor, str):
            continue
        match = PARCELA_REGEX.search(valor)
        if not match:
            continue
        atual = int(match.group(1))
        total = int(match.group(2))
        if atual == total:
            ws.row_dimensions[row].hidden = True
            linhas_quitadas.append(row)
            continue
        cell.value = f"({atual}/{total + 1})"

    wb_rel = _gerar_relatorio(ws, linhas_quitadas, mes, ano)

    pasta = os.path.dirname(caminho_entrada)
    nome_base = os.path.splitext(os.path.basename(caminho_entrada))[0]
    saida = os.path.join(pasta, f"{nome_base}_ATUALIZADO_{mes}_{ano}.xlsx")
    relatorio = os.path.join(pasta, f"{nome_base}_QUITADOS_{mes}_{ano}.xlsx")

    wb.save(saida)
    wb_rel.save(relatorio)
    return saida, relatorio, len(linhas_quitadas)


def processar_planilha_coluna_f(caminho_entrada, mes, ano):
    wb = load_workbook(caminho_entrada)
    ws = wb.active
    linhas_quitadas = []

    for row in range(2, ws.max_row + 1):
        cell_d = ws[f"D{row}"]
        cell_f = ws[f"F{row}"]
        valor_d = cell_d.value
        valor_f = cell_f.value
        if valor_d is None or valor_d == "" or valor_f is None or valor_f == "":
            continue
        try:
            if isinstance(valor_d, str):
                valor_d = re.sub(r'[^\d]', '', valor_d)
            parcela_atual = int(valor_d) if valor_d else 0
            if isinstance(valor_f, str):
                valor_f = re.sub(r'[^\d]', '', valor_f)
            total_parcelas = int(valor_f) if valor_f else 0
            if parcela_atual == 0 or total_parcelas == 0:
                continue
        except (ValueError, TypeError):
            continue

        novo_total = total_parcelas + 1
        ws[f"F{row}"] = novo_total

        if novo_total == parcela_atual:
            ws.row_dimensions[row].hidden = True
            linhas_quitadas.append(row)

    wb_rel = _gerar_relatorio(ws, linhas_quitadas, mes, ano)

    pasta = os.path.dirname(caminho_entrada)
    nome_base = os.path.splitext(os.path.basename(caminho_entrada))[0]
    saida = os.path.join(pasta, f"{nome_base}_ATUALIZADO_{mes}_{ano}.xlsx")
    relatorio = os.path.join(pasta, f"{nome_base}_QUITADOS_{mes}_{ano}.xlsx")

    wb.save(saida)
    wb_rel.save(relatorio)
    return saida, relatorio, len(linhas_quitadas)


# ═══════════════════════════════════════════
# FUNÇÕES - CONTA AZUL (FLUXO DE CAIXA)
# ═══════════════════════════════════════════

def _parse_valor_br(valor_str):
    """Converte valor no formato brasileiro '1.234,56' para float."""
    if valor_str is None:
        return 0.0
    s = str(valor_str).strip()
    if s == "" or s == "0" or s == "-":
        return 0.0
    s = s.replace(".", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


def _ler_fluxo_csv(arquivo_bytes, nome_arquivo):
    """
    Lê um CSV de fluxo de caixa do Conta Azul e retorna dict {dia: (recebimentos, pagamentos)}.
    """
    conteudo = None
    for enc in ['utf-8', 'latin1', 'cp1252', 'iso-8859-1']:
        try:
            conteudo = arquivo_bytes.decode(enc)
            break
        except (UnicodeDecodeError, AttributeError):
            continue

    if conteudo is None:
        raise ValueError(f"Não foi possível decodificar o arquivo {nome_arquivo}")

    linhas = conteudo.replace('\r\n', '\n').replace('\r', '\n').strip().split('\n')
    if not linhas:
        raise ValueError(f"Arquivo {nome_arquivo} está vazio")

    header = linhas[0]
    sep = ';' if header.count(';') >= header.count(',') else ','

    colunas = [c.strip().strip('"') for c in header.split(sep)]

    idx_data = None
    idx_recebimentos = None
    idx_pagamentos = None

    for i, col in enumerate(colunas):
        col_norm = col.lower().strip()
        if idx_data is None and 'data' in col_norm:
            idx_data = i
        if idx_recebimentos is None and 'recebimento' in col_norm:
            idx_recebimentos = i
        if idx_pagamentos is None and 'pagamento' in col_norm:
            idx_pagamentos = i

    if idx_data is None or idx_recebimentos is None or idx_pagamentos is None:
        raise ValueError(
            f"Arquivo '{nome_arquivo}': colunas obrigatórias não encontradas.\n"
            f"Esperado: Data, Recebimentos, Pagamentos.\n"
            f"Encontrado: {colunas}"
        )

    dados_por_dia = {}

    for linha in linhas[1:]:
        if not linha.strip():
            continue

        campos = [c.strip().strip('"') for c in linha.split(sep)]
        if len(campos) <= max(idx_data, idx_recebimentos, idx_pagamentos):
            continue

        data_str = campos[idx_data].strip()
        receb_str = campos[idx_recebimentos].strip()
        pagam_str = campos[idx_pagamentos].strip()

        try:
            partes = data_str.split('/')
            dia = int(partes[0])
            if dia < 1 or dia > 31:
                continue
        except (ValueError, IndexError):
            continue

        receb = _parse_valor_br(receb_str)
        pagam = _parse_valor_br(pagam_str)

        if dia in dados_por_dia:
            r_ant, p_ant = dados_por_dia[dia]
            dados_por_dia[dia] = (r_ant + receb, p_ant + pagam)
        else:
            dados_por_dia[dia] = (receb, pagam)

    return dados_por_dia


def _copiar_estilo(origem, destino):
    """Copia estilo (font, fill, alignment, border, number_format) de uma célula para outra."""
    if origem.font:
        destino.font = copy(origem.font)
    if origem.fill:
        destino.fill = copy(origem.fill)
    if origem.alignment:
        destino.alignment = copy(origem.alignment)
    if origem.border:
        destino.border = copy(origem.border)
    if origem.number_format:
        destino.number_format = origem.number_format


def processar_conta_azul(caminho_receitas, arquivos_fluxo, mes_selecionado, ano_selecionado):
    """
    Consolida os CSVs de fluxo de caixa do Conta Azul e preenche a planilha de Receitas.

    Comportamento:
      - Cria apenas uma linha por dia que possui valores nos CSVs (receita ou despesa > 0)
      - Mantém a mesclagem B1:E1 do título exatamente como no template
      - Linha de TOTAL com mesclagem dupla (2 linhas) igual ao template
      - 4 linhas de assinatura preservadas do template
    """
    from copy import copy as _copy

    # ── 1. Consolidar todos os CSVs ──────────────────────────────────────
    dados_consolidados = {}  # {dia: (total_receb, total_pagam)}

    for nome, conteudo_bytes in arquivos_fluxo:
        dados = _ler_fluxo_csv(conteudo_bytes, nome)
        for dia, (receb, pagam) in dados.items():
            if dia in dados_consolidados:
                r_ant, p_ant = dados_consolidados[dia]
                dados_consolidados[dia] = (r_ant + receb, p_ant + pagam)
            else:
                dados_consolidados[dia] = (receb, pagam)

    # Filtrar apenas dias que têm pelo menos receita ou despesa > 0
    dias_com_dados = sorted(
        dia for dia, (r, p) in dados_consolidados.items() if r > 0 or p > 0
    )

    # ── 2. Abrir planilha de receitas ────────────────────────────────────
    wb = load_workbook(caminho_receitas)

    nome_mes_upper = MESES_PT.get(mes_selecionado, str(mes_selecionado)).upper()
    nome_aba_destino = f"rec desp {mes_selecionado:02d}.{ano_selecionado}"

    # ── 3. Localizar template de formatação ─────────────────────────────
    aba_template = None
    for nome_aba in wb.sheetnames:
        n = nome_aba.strip().lower()
        if n.startswith("rec desp") and nome_aba.strip() != nome_aba_destino:
            aba_template = wb[nome_aba]
            break

    # ── 4. Criar ou limpar a aba de destino ──────────────────────────────
    if nome_aba_destino in wb.sheetnames:
        del wb[nome_aba_destino]

    nome_aba_mes = f"{nome_mes_upper} {ano_selecionado}"
    posicao_insercao = 0
    for i, nome_aba in enumerate(wb.sheetnames):
        if nome_aba.strip().upper() == nome_aba_mes:
            posicao_insercao = i + 1
            break

    ws = wb.create_sheet(title=nome_aba_destino, index=posicao_insercao)

    # ── 5. Copiar dimensões de colunas e linhas do template ───────────────
    if aba_template:
        for col_letter, col_dim in aba_template.column_dimensions.items():
            ws.column_dimensions[col_letter].width = col_dim.width
        for row_num in range(1, 42):
            rd_orig = aba_template.row_dimensions[row_num]
            if rd_orig.height:
                ws.row_dimensions[row_num].height = rd_orig.height
    else:
        ws.column_dimensions['A'].width = 5.8
        ws.column_dimensions['B'].width = 18.7
        ws.column_dimensions['C'].width = 20.0
        ws.column_dimensions['D'].width = 20.2
        ws.column_dimensions['E'].width = 18.5
        ws.row_dimensions[1].height = 38.5
        ws.row_dimensions[2].height = 21.5
        ws.row_dimensions[3].height = 4.5
        ws.row_dimensions[4].height = 57.75

    # ── 6. Linha 1: título mesclado B1:E1 (idêntico ao template) ─────────
    ws.merge_cells('B1:E1')
    cell_titulo = ws.cell(row=1, column=2,
        value="PROVISÃO DE RECEITAS E DESPESAS DIÁRIAS\n MV CONTABILIDADE")
    if aba_template:
        _copiar_estilo(aba_template.cell(row=1, column=2), cell_titulo)
    else:
        cell_titulo.font = Font(bold=True, size=12)
        cell_titulo.fill = PatternFill("solid", fgColor="DAEEF3")
        cell_titulo.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")

    # ── 7. Linha 2: mês e ano ─────────────────────────────────────────────
    cell_mes = ws.cell(row=2, column=3, value=f"{nome_mes_upper} ")
    cell_ano = ws.cell(row=2, column=4, value=ano_selecionado)
    if aba_template:
        _copiar_estilo(aba_template.cell(row=2, column=3), cell_mes)
        _copiar_estilo(aba_template.cell(row=2, column=4), cell_ano)
    else:
        for c in [cell_mes, cell_ano]:
            c.font = Font(bold=True)
            c.fill = PatternFill("solid", fgColor="DAEEF3")
            c.alignment = Alignment(horizontal="center", vertical="center")

    # Linha 3: espaço vazio
    ws.row_dimensions[3].height = 4.5

    # ── 8. Linha 4: headers ───────────────────────────────────────────────
    headers = {2: "DATA DE VENCIMENTO", 3: "RECEITAS", 4: "DESPESAS", 5: "SALDO"}
    for col, texto in headers.items():
        cell = ws.cell(row=4, column=col, value=texto)
        if aba_template:
            _copiar_estilo(aba_template.cell(row=4, column=col), cell)
        else:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="F0F8FA")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # ── 9. Linhas de dados: apenas dias com valores nos CSVs ─────────────
    _fmt_data   = 'DD/MM/YYYY'
    _fmt_valor  = '#,##0.00'
    _align_valor = Alignment(horizontal="right", vertical="center")

    # Estilos de referência da linha de dados do template (linha 5)
    _font_data  = _copy(aba_template.cell(row=5, column=2).font) if aba_template else Font(size=10)
    _align_data = _copy(aba_template.cell(row=5, column=2).alignment) if aba_template \
                  else Alignment(horizontal="center", vertical="center")

    total_receitas = 0.0
    total_despesas = 0.0
    primeira_linha_dados = 5
    linha_atual = primeira_linha_dados

    for dia in dias_com_dados:
        receb, pagam = dados_consolidados[dia]

        # Validar que o dia existe no mês selecionado
        dias_no_mes = calendar.monthrange(ano_selecionado, mes_selecionado)[1]
        if dia > dias_no_mes:
            continue

        nova_data = datetime(ano_selecionado, mes_selecionado, dia)

        # Copiar estilo base do template (linha 5 como referência)
        if aba_template:
            for col in [2, 3, 4, 5]:
                src = aba_template.cell(row=5, column=col)
                dst = ws.cell(row=linha_atual, column=col)
                if src.font:
                    dst.font = _copy(src.font)
                if src.fill and src.fill.patternType:
                    dst.fill = _copy(src.fill)
                if src.border:
                    dst.border = _copy(src.border)

        # Data
        cell_data = ws.cell(row=linha_atual, column=2, value=nova_data)
        cell_data.number_format = _fmt_data
        cell_data.font = _font_data
        cell_data.alignment = _align_data

        # Receitas
        cell_rec = ws.cell(row=linha_atual, column=3, value=receb)
        cell_rec.number_format = _fmt_valor
        cell_rec.alignment = _align_valor

        # Despesas
        cell_desp = ws.cell(row=linha_atual, column=4, value=pagam)
        cell_desp.number_format = _fmt_valor
        cell_desp.alignment = _align_valor

        # Saldo (fórmula =C-D)
        cell_saldo = ws.cell(row=linha_atual, column=5, value=f"=C{linha_atual}-D{linha_atual}")
        cell_saldo.number_format = _fmt_valor
        cell_saldo.alignment = _align_valor

        ws.row_dimensions[linha_atual].height = 14.5

        total_receitas += receb
        total_despesas += pagam
        linha_atual += 1

    ultima_linha_dados = linha_atual - 1  # última linha que tem dado

    # ── 10. Linha de TOTAL — mesclada em 2 linhas (B:E), igual ao template ─
    linha_total = linha_atual  # primeira linha do bloco total
    linha_total2 = linha_atual + 1  # segunda linha do bloco (mesclagem)

    # Mesclar cada coluna verticalmente (2 linhas) como no template: B35:B36 etc.
    for col_letra in ['B', 'C', 'D', 'E']:
        ws.merge_cells(f"{col_letra}{linha_total}:{col_letra}{linha_total2}")

    # Se não há dados, fórmula SUM vazia
    ref_inicio = primeira_linha_dados
    ref_fim = ultima_linha_dados if ultima_linha_dados >= primeira_linha_dados else primeira_linha_dados

    cell_lbl   = ws.cell(row=linha_total, column=2, value="TOTAL")
    cell_sum_c = ws.cell(row=linha_total, column=3, value=f"=SUM(C{ref_inicio}:C{ref_fim})")
    cell_sum_d = ws.cell(row=linha_total, column=4, value=f"=SUM(D{ref_inicio}:D{ref_fim})")
    cell_sum_e = ws.cell(row=linha_total, column=5, value=f"=SUM(E{ref_inicio}:E{ref_fim})")

    for cell in [cell_lbl, cell_sum_c, cell_sum_d, cell_sum_e]:
        cell.number_format = _fmt_valor
    if aba_template:
        for col, cell in zip([2, 3, 4, 5], [cell_lbl, cell_sum_c, cell_sum_d, cell_sum_e]):
            src = aba_template.cell(row=35, column=col)
            if src.font:
                cell.font = _copy(src.font)
            if src.fill and src.fill.patternType:
                cell.fill = _copy(src.fill)
            if src.border:
                cell.border = _copy(src.border)
            if src.alignment:
                cell.alignment = _copy(src.alignment)
        # Reaplica valores após cópia de estilo
        ws.cell(row=linha_total, column=2).value = "TOTAL"
        ws.cell(row=linha_total, column=3).value = f"=SUM(C{ref_inicio}:C{ref_fim})"
        ws.cell(row=linha_total, column=4).value = f"=SUM(D{ref_inicio}:D{ref_fim})"
        ws.cell(row=linha_total, column=5).value = f"=SUM(E{ref_inicio}:E{ref_fim})"
    else:
        for cell in [cell_lbl, cell_sum_c, cell_sum_d, cell_sum_e]:
            cell.font = Font(bold=True)
    ws.row_dimensions[linha_total].height = 15.0
    ws.row_dimensions[linha_total2].height = 15.0

    # ── 11. Assinaturas: 4 linhas após o bloco de TOTAL ──────────────────
    # No template ficam nas linhas 36–39 (4 linhas abaixo de 35).
    # Aqui posicionamos nas 4 linhas seguintes ao bloco de total.
    offset_assinatura = linha_total2 + 1  # começa 1 linha após o fim do total

    if aba_template:
        # Copiar as 4 linhas de assinatura do template (linhas 36–39)
        for i in range(4):
            linha_orig = 36 + i
            linha_dest = offset_assinatura + i
            for col in range(1, 7):
                src = aba_template.cell(row=linha_orig, column=col)
                dst = ws.cell(row=linha_dest, column=col)
                if src.value is not None:
                    dst.value = src.value
                if src.font:
                    dst.font = _copy(src.font)
                if src.fill and src.fill.patternType:
                    dst.fill = _copy(src.fill)
                if src.border:
                    dst.border = _copy(src.border)
                if src.alignment:
                    dst.alignment = _copy(src.alignment)
            # Copiar altura da linha
            rd = aba_template.row_dimensions[linha_orig]
            if rd.height:
                ws.row_dimensions[linha_dest].height = rd.height

    # ── 12. Salvar ────────────────────────────────────────────────────────
    pasta = os.path.dirname(caminho_receitas)
    nome_base = os.path.splitext(os.path.basename(caminho_receitas))[0]
    nome_mes_arquivo = MESES_PT.get(mes_selecionado, str(mes_selecionado))
    saida = os.path.join(pasta, f"{nome_base}_{nome_mes_arquivo}_{ano_selecionado}.xlsx")
    wb.save(saida)

    return saida, len(dias_com_dados), total_receitas, total_despesas


# ═══════════════════════════════════════════
# COMPONENTE DE PROCESSAMENTO (PARCELAS)
# ═══════════════════════════════════════════

def bloco_processamento(key_prefix, descricao, funcao_processar):
    arquivo = st.file_uploader(
        "Arraste o arquivo .xlsx aqui",
        type=['xlsx'],
        key=f"{key_prefix}_uploader",
        help="Apenas arquivos .xlsx são suportados"
    )

    if not arquivo:
        st.markdown(
            f'<p style="text-align:center; color: var(--text-muted); padding: 1.5rem 0; opacity: 0.6;">'
            f'{descricao}</p>',
            unsafe_allow_html=True,
        )
        return

    temp_path = os.path.join(tempfile.gettempdir(), f"mv_{key_prefix}_{arquivo.name}")
    with open(temp_path, "wb") as f:
        f.write(arquivo.getbuffer())

    st.markdown(
        f'<div style="display:flex;align-items:center;gap:8px;padding:8px 14px;'
        f'background:rgba(16,185,129,0.08);border-radius:10px;margin:0.5rem 0 1rem;">'
        f'<span style="font-size:1.15rem;">📎</span>'
        f'<span style="color:#059669;font-weight:500;font-size:0.9rem;">{arquivo.name}</span>'
        f'</div>',
        unsafe_allow_html=True,
    )

    col1, col2 = st.columns(2)
    with col1:
        mes = st.number_input(
            "Mês de referência", min_value=1, max_value=12,
            value=datetime.now().month, step=1, key=f"{key_prefix}_mes"
        )
    with col2:
        ano = st.number_input(
            "Ano de referência", min_value=2020, max_value=2035,
            value=datetime.now().year, step=1, key=f"{key_prefix}_ano"
        )

    st.markdown("<div style='height:0.5rem'></div>", unsafe_allow_html=True)

    if st.button("Processar", key=f"{key_prefix}_btn", use_container_width=True, type="primary"):
        with st.spinner("Processando…"):
            try:
                saida, relatorio, qtd = funcao_processar(temp_path, str(mes), str(ano))

                st.markdown("---")
                st.markdown("#### Resultado")

                m1, m2 = st.columns(2)
                m1.metric("Parcelas quitadas", qtd)
                m2.metric("Arquivos gerados", 2)

                st.markdown("<div style='height:0.5rem'></div>", unsafe_allow_html=True)

                d1, d2 = st.columns(2)
                with d1:
                    with open(saida, "rb") as f:
                        st.download_button(
                            "⬇  Planilha Atualizada", f,
                            file_name=os.path.basename(saida),
                            use_container_width=True,
                        )
                with d2:
                    with open(relatorio, "rb") as f:
                        st.download_button(
                            "⬇  Relatório de Quitados", f,
                            file_name=os.path.basename(relatorio),
                            use_container_width=True,
                        )

            except Exception as e:
                st.error(f"Erro ao processar: {str(e)}")
            finally:
                if os.path.exists(temp_path):
                    os.remove(temp_path)


# ═══════════════════════════════════════════
# COMPONENTE CONTA AZUL
# ═══════════════════════════════════════════

def bloco_conta_azul():
    st.markdown(
        '<p style="font-size:0.88rem; color: var(--mv-muted, #64748b); margin-bottom: 1rem;">'
        'Anexe os <strong>CSVs de Fluxo de Caixa Diário</strong> exportados do Conta Azul e a '
        '<strong>planilha de Receitas (.xlsx)</strong>. Os <em>Recebimentos</em> de cada dia '
        '(col B do CSV) serão somados e inseridos em <strong>Receitas</strong> (col C), e os '
        '<em>Pagamentos</em> (col C do CSV) em <strong>Despesas</strong> (col D) da planilha. '
        'Apenas dias com valores são incluídos.</p>',
        unsafe_allow_html=True,
    )

    arquivos_fluxo = st.file_uploader(
        "📂 Planilhas de Fluxo de Caixa Diário (.csv)",
        type=['csv'],
        key="ca_fluxo_uploader",
        accept_multiple_files=True,
        help="Exporte do Conta Azul: Relatórios → Fluxo de Caixa Diário → CSV"
    )

    arquivo_receitas = st.file_uploader(
        "📋 Planilha de Receitas e Despesas (.xlsx)",
        type=['xlsx'],
        key="ca_receitas_uploader",
        help="A planilha com abas mensais (ex: '2-_Receitas_MV.xlsx'). "
             "Os dados serão inseridos em uma nova aba 'rec desp MM.AAAA'."
    )

    if arquivos_fluxo:
        for arq in arquivos_fluxo:
            st.markdown(
                f'<div style="display:flex;align-items:center;gap:8px;padding:6px 14px;'
                f'background:rgba(59,130,246,0.06);border-radius:10px;margin:0.2rem 0;">'
                f'<span style="font-size:1rem;">📄</span>'
                f'<span style="color:#2563eb;font-weight:500;font-size:0.85rem;">{arq.name}</span>'
                f'</div>',
                unsafe_allow_html=True,
            )

    if arquivo_receitas:
        st.markdown(
            f'<div style="display:flex;align-items:center;gap:8px;padding:8px 14px;'
            f'background:rgba(16,185,129,0.08);border-radius:10px;margin:0.3rem 0;">'
            f'<span style="font-size:1.15rem;">📎</span>'
            f'<span style="color:#059669;font-weight:500;font-size:0.9rem;">{arquivo_receitas.name}</span>'
            f'</div>',
            unsafe_allow_html=True,
        )

    if not arquivos_fluxo or not arquivo_receitas:
        st.markdown(
            '<p style="text-align:center; color: var(--text-muted); padding: 1.5rem 0; opacity: 0.6;">'
            'Anexe pelo menos um CSV de Fluxo de Caixa e a planilha de Receitas para continuar</p>',
            unsafe_allow_html=True,
        )
        return

    st.markdown("<div style='height:0.5rem'></div>", unsafe_allow_html=True)
    col1, col2 = st.columns(2)
    with col1:
        mes = st.number_input(
            "Mês de referência", min_value=1, max_value=12,
            value=datetime.now().month, step=1, key="ca_mes"
        )
    with col2:
        ano = st.number_input(
            "Ano de referência", min_value=2020, max_value=2035,
            value=datetime.now().year, step=1, key="ca_ano"
        )

    nome_aba_preview = f"rec desp {mes:02d}.{ano}"
    st.markdown(
        f'<div style="background:rgba(59,130,246,0.06);border:1px solid rgba(59,130,246,0.15);'
        f'border-radius:10px;padding:10px 16px;margin:0.8rem 0;">'
        f'<span style="color:#64748b;font-size:0.82rem;">Aba que será criada/atualizada: </span>'
        f'<strong style="color:#1e40af;">{nome_aba_preview}</strong>'
        f'</div>',
        unsafe_allow_html=True,
    )

    st.markdown("<div style='height:0.3rem'></div>", unsafe_allow_html=True)

    if st.button("Processar Fluxo de Caixa", key="ca_btn", use_container_width=True, type="primary"):
        with st.spinner("Consolidando e preenchendo a planilha…"):
            temp_receitas = None
            try:
                temp_receitas = os.path.join(
                    tempfile.gettempdir(), f"mv_ca_{arquivo_receitas.name}"
                )
                with open(temp_receitas, "wb") as f:
                    f.write(arquivo_receitas.getbuffer())

                lista_fluxos = [(arq.name, arq.getvalue()) for arq in arquivos_fluxo]

                saida, dias_processados, total_rec, total_desp = processar_conta_azul(
                    temp_receitas, lista_fluxos, int(mes), int(ano)
                )

                st.markdown("---")
                st.markdown("#### Resultado")

                m1, m2, m3 = st.columns(3)
                m1.metric("Dias com dados", dias_processados)
                m2.metric("Total Receitas", f"R$ {total_rec:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))
                m3.metric("Total Despesas", f"R$ {total_desp:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))

                st.markdown("<div style='height:0.5rem'></div>", unsafe_allow_html=True)

                nome_mes_label = MESES_PT.get(int(mes), str(mes))
                st.markdown(
                    f'<div style="background:rgba(59,130,246,0.06);border:1px solid rgba(59,130,246,0.15);'
                    f'border-radius:10px;padding:12px 16px;margin:0.5rem 0;">'
                    f'<span style="font-weight:600;color:#1e40af;">📊 {nome_mes_label}/{ano}</span>'
                    f'<span style="color:#64748b;font-size:0.85rem;"> — '
                    f'{len(arquivos_fluxo)} arquivo(s) de fluxo consolidado(s) | '
                    f'Aba <strong>{nome_aba_preview}</strong> criada</span>'
                    f'</div>',
                    unsafe_allow_html=True,
                )

                with open(saida, "rb") as f:
                    st.download_button(
                        "⬇  Baixar Planilha de Receitas Atualizada",
                        f,
                        file_name=os.path.basename(saida),
                        use_container_width=True,
                    )

            except Exception as e:
                st.error(f"Erro ao processar: {str(e)}")
            finally:
                if temp_receitas and os.path.exists(temp_receitas):
                    os.remove(temp_receitas)


# ═══════════════════════════════════════════
# INTERFACE PRINCIPAL (PÓS-LOGIN)
# ═══════════════════════════════════════════

def tela_principal():
    st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=DM+Sans:ital,opsz,wght@0,9..40,300;0,9..40,500;0,9..40,700;1,9..40,400&display=swap');

    :root {
        --mv-navy:    #0f1b2d;
        --mv-blue:    #1b3a5c;
        --mv-accent:  #3b82f6;
        --mv-accent2: #60a5fa;
        --mv-surface: #f8fafc;
        --mv-border:  #e2e8f0;
        --mv-text:    #1e293b;
        --mv-muted:   #64748b;
        --mv-radius:  14px;
    }

    .stApp { background: var(--mv-surface) !important; }
    html, body, .stApp, .stApp * { font-family: 'DM Sans', sans-serif !important; }

    .mv-header {
        background: linear-gradient(135deg, var(--mv-navy) 0%, var(--mv-blue) 60%, #234e7a 100%);
        padding: 2.2rem 2rem 1.8rem;
        border-radius: 0 0 var(--mv-radius) var(--mv-radius);
        margin: -1rem -1rem 1.8rem;
        position: relative; overflow: hidden;
    }
    .mv-header::before {
        content: ''; position: absolute; top: -40%; right: -10%;
        width: 320px; height: 320px;
        background: radial-gradient(circle, rgba(59,130,246,0.12) 0%, transparent 70%);
        border-radius: 50%;
    }
    .mv-header h1 {
        color: #fff; font-size: 1.6rem; font-weight: 700;
        letter-spacing: 2.5px; margin: 0; position: relative;
    }
    .mv-header p {
        color: rgba(255,255,255,0.55); font-size: 0.82rem;
        font-weight: 400; margin: 0.35rem 0 0;
        letter-spacing: 0.5px; position: relative;
    }
    .mv-header .user-badge {
        position: absolute; top: 50%; right: 2rem;
        transform: translateY(-50%);
        display: flex; align-items: center; gap: 10px;
        color: rgba(255,255,255,0.7); font-size: 0.82rem;
    }
    .mv-header .user-badge .avatar {
        width: 34px; height: 34px;
        background: rgba(255,255,255,0.15);
        border-radius: 10px;
        display: flex; align-items: center; justify-content: center;
        font-size: 0.85rem;
    }

    .stTabs [data-baseweb="tab-list"] {
        gap: 0; background: #fff; border-radius: var(--mv-radius);
        padding: 4px; border: 1px solid var(--mv-border);
        box-shadow: 0 1px 3px rgba(0,0,0,0.04);
    }
    .stTabs [data-baseweb="tab"] {
        border-radius: 10px; padding: 0.55rem 1.2rem;
        font-size: 0.85rem; font-weight: 500;
        color: var(--mv-muted); background: transparent; white-space: nowrap;
    }
    .stTabs [aria-selected="true"] {
        background: var(--mv-navy) !important; color: #fff !important; font-weight: 600;
    }
    .stTabs [data-baseweb="tab-border"],
    .stTabs [data-baseweb="tab-highlight"] { display: none; }

    [data-testid="stFileUploader"] { border-radius: var(--mv-radius); }
    [data-testid="stFileUploader"] section {
        border-radius: var(--mv-radius) !important;
        border: 2px dashed var(--mv-border) !important;
        padding: 1.5rem !important; transition: border-color 0.2s;
    }
    [data-testid="stFileUploader"] section:hover { border-color: var(--mv-accent) !important; }
    [data-testid="stFileUploader"] svg { display: none; }

    .stButton > button[kind="primary"],
    button[data-testid="stBaseButton-primary"] {
        background: var(--mv-navy) !important; color: #fff !important;
        border: none !important; border-radius: 10px !important;
        padding: 0.65rem 1.5rem !important; font-weight: 600 !important;
        font-size: 0.9rem !important; letter-spacing: 0.3px; transition: all 0.2s ease;
    }
    .stButton > button[kind="primary"]:hover,
    button[data-testid="stBaseButton-primary"]:hover {
        background: var(--mv-blue) !important;
        box-shadow: 0 4px 14px rgba(15,27,45,0.25); transform: translateY(-1px);
    }

    .stDownloadButton > button {
        background: #fff !important; color: var(--mv-navy) !important;
        border: 1.5px solid var(--mv-border) !important; border-radius: 10px !important;
        font-weight: 600 !important; font-size: 0.85rem !important; transition: all 0.2s ease;
    }
    .stDownloadButton > button:hover {
        border-color: var(--mv-accent) !important; color: var(--mv-accent) !important;
        box-shadow: 0 2px 8px rgba(59,130,246,0.12);
    }

    [data-testid="stMetric"] {
        background: #fff; border: 1px solid var(--mv-border);
        border-radius: 12px; padding: 1rem 1.2rem;
        box-shadow: 0 1px 3px rgba(0,0,0,0.04);
    }
    [data-testid="stMetricValue"] { color: var(--mv-navy); font-weight: 700; }
    [data-testid="stMetricLabel"] { color: var(--mv-muted); font-size: 0.8rem; font-weight: 500; }

    [data-testid="stNumberInput"] label {
        font-size: 0.85rem; font-weight: 500; color: var(--mv-text);
    }

    hr { border-color: var(--mv-border) !important; opacity: 0.5; }
    #MainMenu, footer, header { visibility: hidden; }
    </style>
    """, unsafe_allow_html=True)

    nome = st.session_state.get("nome_usuario", "Usuário")
    iniciais = "".join([p[0].upper() for p in nome.split()[:2]]) if nome else "U"

    st.markdown(f"""
        <div class="mv-header">
            <h1>MV TEC</h1>
            <p>Gestão de parcelas e empréstimos</p>
            <div class="user-badge">
                <span>{nome}</span>
                <div class="avatar">{iniciais}</div>
            </div>
        </div>
    """, unsafe_allow_html=True)

    with st.sidebar:
        st.markdown(f"**Logado como:** {nome}")
        st.markdown(f"**Perfil:** {st.session_state.get('perfil', 'operador').title()}")
        st.markdown("---")
        if st.button("🚪 Sair", use_container_width=True):
            for key in ["autenticado", "usuario", "nome_usuario", "perfil"]:
                st.session_state.pop(key, None)
            st.rerun()

    tab1, tab2, tab3 = st.tabs(["Empréstimos Individuais", "Empréstimos", "Conta Azul"])

    with tab1:
        bloco_processamento(
            key_prefix="ind",
            descricao="Processa parcelas no formato (X/Y) na coluna H",
            funcao_processar=processar_planilha,
        )

    with tab2:
        bloco_processamento(
            key_prefix="col",
            descricao="Incrementa coluna F — se F = D após incremento, marca como quitado",
            funcao_processar=processar_planilha_coluna_f,
        )

    with tab3:
        bloco_conta_azul()


# ═══════════════════════════════════════════
# PONTO DE ENTRADA
# ═══════════════════════════════════════════

def main():
    st.set_page_config(
        page_title="MV TEC",
        page_icon="📊",
        layout="centered",
        initial_sidebar_state="collapsed",
    )

    if st.session_state.get("autenticado"):
        tela_principal()
    else:
        tela_login()


if __name__ == "__main__":
    main()
